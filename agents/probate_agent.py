from __future__ import annotations
"""
agents/probate_agent.py — Scrapes probate court filings for Ohio POC counties.

County portals (all public, all JS-rendered — requires Playwright):
    Cuyahoga: https://probate.cuyahogacounty.gov/pa/  (PROWARE ASP.NET — TOS gate on first visit)
    Lake:     https://courtrecords.lakecountyclerk.org (ShowCase)
    Mahoning: https://probate.mahoningcountyoh.gov     (CivicPlus)

Each county has its own search flow — scrapers are implemented separately
and registered in COUNTY_SCRAPERS. The 9-step agent pipeline is shared.

CLI:
    python agents/probate_agent.py --county cuyahoga --state OH
    python agents/probate_agent.py --all-counties --state OH
    python agents/probate_agent.py --county cuyahoga --state OH --days 7
"""

import argparse
import asyncio
import os
from datetime import date, timedelta
from typing import Optional

from dotenv import load_dotenv
from playwright.async_api import Page, async_playwright

from db.client import get_client, insert_row, update_row
from enrichment.waterfall import enrich_lead
from routing.notify import send_sms
from routing.va_router import route_lead
from scoring.score import score_lead
from utils.deduper import is_duplicate
from utils.logger import get_logger
from verification.verify_leads import verify_raw_record

load_dotenv()

log = get_logger("probate_agent")

OHIO_COUNTIES = ["cuyahoga", "lake", "mahoning"]
DEFAULT_LOOKBACK_DAYS = 3  # fetch filings from the last N days on each run

# Mahoning County reCAPTCHA v2 site key (confirmed 2026-04-18 — static on Tyler Odyssey portal)
_MAHONING_RECAPTCHA_SITEKEY = "6LfqmHkUAAAAAAKhHRHuxUy6LOMRZSG2LvSwWPO9"


def _brightdata_proxy_config() -> Optional[dict]:
    """Return a Playwright proxy dict for BrightData residential proxy, or None if unconfigured.

    Credentials must be set in .env:
      BRIGHTDATA_HOST       (default: brd.superproxy.io)
      BRIGHTDATA_PORT       (default: 33335)
      BRIGHTDATA_USERNAME   customer-XXXX-zone-YYYY format from BrightData dashboard
      BRIGHTDATA_PASSWORD
    """
    host = os.environ.get("BRIGHTDATA_HOST", "brd.superproxy.io")
    port = os.environ.get("BRIGHTDATA_PORT", "33335")
    username = os.environ.get("BRIGHTDATA_USERNAME", "")
    password = os.environ.get("BRIGHTDATA_PASSWORD", "")
    if not username or not password:
        return None
    return {
        "server": f"http://{host}:{port}",
        "username": username,
        "password": password,
    }


async def _twocaptcha_solve_recaptcha(sitekey: str, page_url: str) -> str:
    """Solve a reCAPTCHA v2 challenge via 2captcha API.

    Runs the blocking 2captcha SDK call in a thread executor so it does not
    block the asyncio event loop. Typical solve time: 15–45 seconds.
    Raises RuntimeError if TWOCAPTCHA_API_KEY is missing or the solve fails.
    """
    api_key = os.environ.get("TWOCAPTCHA_API_KEY", "")
    if not api_key:
        raise RuntimeError(
            "TWOCAPTCHA_API_KEY is not set in .env — required for Mahoning County."
        )

    def _solve() -> str:
        from twocaptcha import TwoCaptcha
        solver = TwoCaptcha(api_key)
        result = solver.recaptcha(sitekey=sitekey, url=page_url)
        return result["code"]

    loop = asyncio.get_event_loop()
    token = await loop.run_in_executor(None, _solve)
    log.debug(f"2captcha solve complete — token length: {len(token)}")
    return token


# =============================================================================
# Cuyahoga County — probate.cuyahogacounty.gov/pa/
# PROWARE ASP.NET system. Requires TOS acceptance on first visit (session cookie).
# No date-range filter — search by last name prefix + ESTATE category + year.
# Results: Name, Case Number, Address, Role, Alias.
# Case detail (CaseSummary.aspx) has filing date; only fetched for new case numbers.
# =============================================================================

# Common 2-letter surname prefixes covering ~80% of US surnames.
# Searched in order; results deduplicated by case number.
_SURNAME_PREFIXES = [
    "AB", "AD", "AL", "AM", "AN", "AR", "AT",
    "BA", "BE", "BI", "BL", "BO", "BR", "BU",
    "CA", "CH", "CI", "CL", "CO", "CR", "CU",
    "DA", "DE", "DI", "DO", "DU",
    "ED", "EL", "EN", "EV",
    "FA", "FE", "FI", "FL", "FO", "FR",
    "GA", "GE", "GI", "GL", "GO", "GR", "GU",
    "HA", "HE", "HI", "HO", "HU",
    "JA", "JE", "JO", "JU",
    "KA", "KE", "KI", "KO",
    "LA", "LE", "LI", "LO", "LU",
    "MA", "MC", "ME", "MI", "MO", "MU",
    "NA", "NE", "NI", "NO",
    "OL", "OR", "OS",
    "PA", "PE", "PH", "PO", "PR",
    "RA", "RE", "RI", "RO", "RU",
    "SA", "SC", "SE", "SH", "SI", "SM", "SN", "SO", "SP", "ST", "SU", "SW",
    "TA", "TE", "TH", "TI", "TO", "TR", "TU",
    "VA", "VE", "VI",
    "WA", "WE", "WH", "WI", "WO",
    "YO", "ZA",
]


async def _cuyahoga_accept_tos(page: Page) -> None:
    """Navigate to the Cuyahoga Probate docket and accept the Terms of Service.

    Sets a session cookie that persists for subsequent navigations within the
    same Playwright context. Safe to call once per session.
    """
    await page.goto("https://probate.cuyahogacounty.gov/pa/", wait_until="networkidle", timeout=30000)
    if await page.query_selector("#mpContentPH_btnYes"):
        async with page.expect_navigation(wait_until="networkidle", timeout=30000):
            await page.click("#mpContentPH_btnYes", no_wait_after=True)


async def _cuyahoga_search_prefix(page: Page, prefix: str, year_str: str) -> list[dict]:
    """Run a single surname-prefix search and collect all DECEDENT rows across pages.

    Returns list of dicts: {case_number, owner_name, address}.
    Does NOT fetch filing dates — that's done in a separate pass.
    """
    SEARCH_URL = "https://probate.cuyahogacounty.gov/pa/CaseSearch.aspx"
    candidates: list[dict] = []
    seen_in_prefix: set[str] = set()

    await page.goto(SEARCH_URL, wait_until="networkidle", timeout=30000)
    await page.select_option("#mpContentPH_ddlPplCaseCat", value="EST")
    await page.fill("#mpContentPH_txtPPLYear", year_str)
    await page.fill("#mpContentPH_txtLName", prefix)

    async with page.expect_navigation(wait_until="networkidle", timeout=45000):
        await page.click("#mpContentPH_btnSearchByPerson", no_wait_after=True)

    page_num = 1
    while True:
        content = await page.inner_text("body")
        if "No results were found" in content:
            break

        rows = await page.query_selector_all("#mpContentPH_gvSearchResults tr")
        for row in rows:
            cells = await row.query_selector_all("td")
            if len(cells) < 4:
                continue
            texts = [await c.inner_text() for c in cells]
            role = texts[3].strip().upper()
            if role != "DECEDENT":
                continue
            case_num = texts[1].strip()
            if case_num in seen_in_prefix:
                continue
            seen_in_prefix.add(case_num)
            candidates.append({
                "case_number": case_num,
                "owner_name": texts[0].strip(),
                "address": texts[2].strip(),
            })

        # Advance to next page if available (page links rendered as plain numbers in a row)
        all_links = await page.query_selector_all("#mpContentPH_gvSearchResults a, tr td a")
        next_page_link = None
        for a in all_links:
            href = await a.get_attribute("href") or ""
            if f"Page${page_num + 1}" in href:
                next_page_link = a
                break
        if not next_page_link:
            break
        await next_page_link.click()
        await page.wait_for_load_state("networkidle")
        page_num += 1

    return candidates


async def _cuyahoga_fetch_filing_date(page: Page, case_num: str, year_str: str) -> Optional[str]:
    """Search for a specific case number and click into its summary to get the filing date.

    Returns ISO date string (YYYY-MM-DD) or None if not found.
    Case number format: 2026EST306311 → txtCaseYear=2026, txtCaseNum=306311
    """
    import re

    SEARCH_URL = "https://probate.cuyahogacounty.gov/pa/CaseSearch.aspx"
    num_part = re.sub(r"^\d{4}EST0*", "", case_num)

    await page.goto(SEARCH_URL, wait_until="networkidle", timeout=30000)
    await page.select_option("#mpContentPH_ddlCaseCat", value="EST")
    await page.fill("#mpContentPH_txtCaseYear", year_str)
    await page.fill("#mpContentPH_txtCaseNum", num_part)

    async with page.expect_navigation(wait_until="networkidle", timeout=45000):
        await page.click("#mpContentPH_btnSearchByCase", no_wait_after=True)

    # The results page shows parties. Click the first name link to reach CaseSummary.aspx.
    links = await page.query_selector_all('a[href*="javascript:__doPostBack"]')
    for link in links:
        href = await link.get_attribute("href") or ""
        if "lbName" in href:
            async with page.expect_navigation(wait_until="networkidle", timeout=45000):
                await link.click(no_wait_after=True)
            break

    content = await page.inner_text("body")
    match = re.search(r"Filing Date:\s+\w+,\s+(\w+ \d+, \d+)", content)
    if match:
        from datetime import datetime as dt
        try:
            return dt.strptime(match.group(1), "%B %d, %Y").date().isoformat()
        except ValueError:
            pass
    return None


async def _fetch_cuyahoga(page: Page, since: date) -> list[dict]:
    """Scrape Cuyahoga County probate (ESTATE) filings since `since` date.

    Portal: https://probate.cuyahogacounty.gov/pa/ (PROWARE ASP.NET)

    Two-pass strategy (portal has no date-range filter):
      Pass 1 — Sweep 113 common 2-letter surname prefixes with ESTATE + year.
               Collect DECEDENT case numbers and addresses. Deduplicate.
      Pass 2 — For each new case number (not already in raw_leads), fetch
               the filing date from the case summary page. Skip cases older
               than `since`. Log and store cases within the lookback window.

    Delay: 3–5 seconds between requests per Zillow TOS crawl-rate guidance.
    Timeout: 45 seconds per request (portal is slow).
    """
    import random

    base_url = "https://probate.cuyahogacounty.gov/pa/"
    results = []
    year_str = str(since.year)

    await _cuyahoga_accept_tos(page)

    # Pass 1: collect all DECEDENT case numbers across surname prefixes
    all_candidates: dict[str, dict] = {}  # case_number → {owner_name, address}
    consecutive_failures = 0
    MAX_CONSECUTIVE_FAILURES = 6  # abort after 6 straight timeouts (~4-5 min wasted max)
    for prefix in _SURNAME_PREFIXES:
        log.debug(f"Cuyahoga pass 1: prefix '{prefix}'")
        try:
            new_rows = await _cuyahoga_search_prefix(page, prefix, year_str)
            consecutive_failures = 0  # reset on success
            for row in new_rows:
                cn = row["case_number"]
                if cn not in all_candidates:
                    all_candidates[cn] = row
        except Exception as e:
            log.warning(f"Cuyahoga: search failed for prefix '{prefix}': {e}")
            consecutive_failures += 1
            if consecutive_failures >= MAX_CONSECUTIVE_FAILURES:
                log.error(
                    f"Cuyahoga probate: {MAX_CONSECUTIVE_FAILURES} consecutive prefix failures — "
                    f"site appears down, aborting prefix scan early with {len(all_candidates)} candidates"
                )
                break
        await asyncio.sleep(random.uniform(3, 5))

    log.info(f"Cuyahoga pass 1 complete: {len(all_candidates)} unique DECEDENT case numbers")

    # Pass 2: fetch filing date for each new case; filter to since window
    from utils.deduper import is_duplicate as _is_dup
    for case_num, info in all_candidates.items():
        # Skip if already in DB (idempotent)
        if _is_dup("cuyahoga", "probate", property_address=info["address"], owner_name=info["owner_name"]):
            log.debug(f"Cuyahoga: skipping duplicate {case_num}")
            continue

        await asyncio.sleep(random.uniform(3, 5))
        try:
            filing_date_str = await _cuyahoga_fetch_filing_date(page, case_num, year_str)
        except Exception as e:
            log.debug(f"Cuyahoga: could not fetch date for {case_num}: {e}")
            filing_date_str = None

        if filing_date_str:
            from datetime import datetime as dt
            filing_date = dt.strptime(filing_date_str, "%Y-%m-%d").date()
            if filing_date < since:
                log.debug(f"Cuyahoga: {case_num} filed {filing_date_str} — older than window, skipping")
                continue

        results.append({
            "case_number": case_num,
            "owner_name": info["owner_name"],
            "filing_date": filing_date_str or "",
            "case_type": "Estate",
            "_county": "Cuyahoga",
            "_source_url": base_url,
            "_address": info["address"],
        })
        log.debug(f"Cuyahoga: queued {case_num} — {info['owner_name']} ({filing_date_str})")

    log.info(f"Cuyahoga: fetched {len(results)} new ESTATE filings since {since}")
    return results


# =============================================================================
# Lake County — phoenix.lakecountyohio.gov/jwprobate/ (Phoenix/NIC system)
#
# BLOCKED: Cloudflare hard-blocks ALL headless requests on this domain.
# Confirmed 2026-04-18 — both portal paths blocked:
#   /probate/home.page.2   (marriage + estate)
#   /jwprobate/home.page   (court record search — confirmed via lakecountyohio.gov links)
#
# BrightData residential proxy (zone: intel_lead, port 33335) is flagged by
# Cloudflare for this domain — returns "403 Attention Required / Sorry, you
# have been blocked". Stealth headers do not help — the block is at the IP
# reputation layer, not the browser fingerprint layer.
#
# Fix required (owner action):
#   Add a BrightData Scraping Browser zone in the BrightData dashboard, then
#   set BRIGHTDATA_SBR_WS in .env:
#     BRIGHTDATA_SBR_WS=wss://<username>:<password>@brd.superproxy.io:9222
#   The scraper will auto-enable once that variable is present.
#   See Option A instructions added to .env.example.
# =============================================================================


class CloudflareBlockedError(RuntimeError):
    """Raised when a portal is hard-blocked by Cloudflare and cannot be scraped."""


async def _fetch_lake(page: Page, since: date) -> list[dict]:
    """Scrape Lake County probate filings since `since` date.

    Portal: https://phoenix.lakecountyohio.gov/jwprobate/home.page
    System: Phoenix (NIC/Tyler Technologies)

    Cloudflare bypass: requires BrightData Scraping Browser (not residential proxy).
    When BRIGHTDATA_SBR_WS is set in .env, a CDP-connected browser is used instead
    of the standard Playwright context — the caller (_run_async) handles the switch.
    This function assumes the page is already routed through the Scraping Browser.

    Raises CloudflareBlockedError if CF block is still detected after navigation,
    which _run_async catches to send an owner SMS and skip Lake County gracefully.
    """
    import random

    search_url = "https://phoenix.lakecountyohio.gov/jwprobate/SearchCases.page"
    results = []

    # --- Navigate and confirm Cloudflare is bypassed ---
    await page.goto(search_url, wait_until="domcontentloaded", timeout=30000)
    await page.wait_for_timeout(3000)
    body_text = (await page.inner_text("body")).lower()
    if any(phrase in body_text for phrase in (
        "just a moment", "cf-mitigated", "sorry, you have been blocked",
        "attention required", "performing security verification",
    )):
        raise CloudflareBlockedError(
            "Lake County: phoenix.lakecountyohio.gov is hard-blocked by Cloudflare. "
            "BrightData Scraping Browser zone required — set BRIGHTDATA_SBR_WS in .env. "
            "See CLAUDE.md Lake County section for setup instructions."
        )

    # --- Fill search form ---
    # Phoenix (NIC) estate case search — these selectors target the standard
    # Phoenix 2.x case-search form used by Ohio county portals.
    since_str = since.strftime("%m/%d/%Y")
    today_str = date.today().strftime("%m/%d/%Y")

    # Case type: select "Estate" or "Probate" from the dropdown if present.
    # Phoenix portals use either a <select> or a text field with autocomplete.
    try:
        ct_select = await page.query_selector(
            "select[id*='CaseType'], select[name*='CaseType'], select[id*='caseType']"
        )
        if ct_select:
            # Try common label values; fall through silently if neither matches.
            for label in ("Estate", "Probate", "ESTATE", "PROBATE"):
                try:
                    await ct_select.select_option(label=label)
                    break
                except Exception:
                    continue
    except Exception as e:
        log.debug(f"Lake County: case-type selector not found or not applicable: {e}")

    # File-date range — Phoenix renders these as text inputs (MM/DD/YYYY).
    # Selector order: prefer id-based, fall back to name-based or placeholder.
    date_from_sel = (
        "input[id*='FileDateFrom'], input[id*='fileDateFrom'], "
        "input[name*='FileDateFrom'], input[placeholder*='From Date']"
    )
    date_to_sel = (
        "input[id*='FileDateTo'], input[id*='fileDateTo'], "
        "input[name*='FileDateTo'], input[placeholder*='To Date']"
    )
    await page.fill(date_from_sel, since_str)
    await page.fill(date_to_sel, today_str)

    await asyncio.sleep(random.uniform(1, 2))
    async with page.expect_navigation(wait_until="networkidle", timeout=45000):
        await page.click(
            "button[type='submit'], input[type='submit'], button[id*='Search'], button[id*='search']"
        )

    # --- Parse paginated results ---
    page_num = 1
    while True:
        body_text = await page.inner_text("body")
        if "no cases found" in body_text.lower() or "no results" in body_text.lower():
            break

        # Phoenix results table: <table> with rows containing case data.
        # Column order (typical): Case Number | Style/Name | File Date | Case Type | Status
        rows = await page.query_selector_all(
            "table#SearchResults tr, table.search-results tr, "
            "table[id*='Results'] tr, table[class*='results'] tr"
        )
        if not rows:
            log.warning(
                "Lake County: results table not found — page structure may have changed. "
                f"Page title: {await page.title()}"
            )
            break

        for row in rows:
            cells = await row.query_selector_all("td")
            if len(cells) < 3:
                continue
            texts = [await c.inner_text() for c in cells]
            # Skip header-looking rows
            if texts[0].strip().lower() in ("case number", "case no", "#"):
                continue

            case_num = texts[0].strip()
            # Style / owner name is typically the second column
            owner = texts[1].strip() if len(texts) > 1 else ""
            # File date is typically the third column
            file_date_raw = texts[2].strip() if len(texts) > 2 else ""
            case_type_raw = texts[3].strip() if len(texts) > 3 else ""

            # Only collect Estate / Probate case types
            if case_type_raw and not any(
                t in case_type_raw.upper() for t in ("ESTATE", "PROBATE", "EST")
            ):
                continue

            results.append({
                "case_number": case_num,
                "owner_name": owner,
                "filing_date": file_date_raw,
                "case_type": case_type_raw or "Estate",
                "_county": "Lake",
                "_source_url": search_url,
                "_address": "",  # not available in Phoenix search results; enrichment fills this
            })

        # Advance to next page if available
        next_link = await page.query_selector(
            f"a[href*='Page${page_num + 1}'], a[title='Next Page'], a[rel='next']"
        )
        if not next_link:
            break
        await next_link.click()
        await page.wait_for_load_state("networkidle")
        await asyncio.sleep(random.uniform(2, 4))
        page_num += 1

    log.info(f"Lake County: fetched {len(results)} probate filings since {since}")
    return results


# =============================================================================
# Mahoning County — portal-ohmahoning.tylertech.cloud/Portal/ (Tyler Odyssey)
#
# BLOCKED: Google reCAPTCHA v2 required on all searches, including
# authenticated sessions (CaptchaDisabledForAuthenticated=False).
# Confirmed 2026-04-18: reCAPTCHA iframe present before form submission.
# Key: 6LfqmHkUAAAAAAKhHRHuxUy6LOMRZSG2LvSwWPO9
#
# Resolution options (owner decision):
#   1. 2captcha.com API — solves reCAPTCHA v2 for ~$2.99/1000; at ~30/day
#      for Mahoning that's < $0.10/day. Integrate via 2captcha Python SDK.
#   2. CapSolver or Anti-Captcha as alternative services
#   3. Evaluate whether Tyler Tech offers a data export / bulk API for
#      registered court users (contact Tyler Tech support)
#
# Direct portal URL: https://portal-ohmahoning.tylertech.cloud/Portal/
# Search URL: https://portal-ohmahoning.tylertech.cloud/Portal/Home/Dashboard/29
# POST endpoint: /Portal/SmartSearch/SmartSearch/SmartSearch
# Has date-range filter: FileDateStart / FileDateEnd (in Advanced Options)
# =============================================================================

async def _fetch_mahoning(page: Page, since: date) -> list[dict]:
    """Scrape Mahoning County probate filings since `since` date.

    Portal: https://portal-ohmahoning.tylertech.cloud/Portal/
    System: Tyler Technologies Odyssey (SmartSearch)

    reCAPTCHA bypass: 2captcha.com API (TWOCAPTCHA_API_KEY in .env).
    The portal presents reCAPTCHA v2 on every search submission — including
    for authenticated sessions (CaptchaDisabledForAuthenticated=False).
    Site key: 6LfqmHkUAAAAAAKhHRHuxUy6LOMRZSG2LvSwWPO9 (static, confirmed 2026-04-18).

    Search flow (Tyler Odyssey SmartSearch):
      1. Load Dashboard/29 (Probate division).
      2. Expand Advanced Options and fill FileDateStart / FileDateEnd.
      3. Set case type filter to Estate via the NodeDesc autocomplete.
      4. Solve reCAPTCHA v2 via 2captcha, inject token into the hidden textarea.
      5. Submit #btnSSSubmit — Odyssey POSTs to SmartSearch/SmartSearch and
         returns JSON with a SearchResults array.
      6. Paginate via the "next" link until all results are collected.
    """
    import json
    import random

    dashboard_url = "https://portal-ohmahoning.tylertech.cloud/Portal/Home/Dashboard/29"
    results = []

    since_str = since.strftime("%m/%d/%Y")
    today_str = date.today().strftime("%m/%d/%Y")

    await page.goto(dashboard_url, wait_until="networkidle", timeout=30000)

    # --- Expand Advanced Options ---
    adv_btn = await page.query_selector("#AdvOptions, [data-toggle='#AdvOptions'], a[href*='AdvOptions']")
    if adv_btn:
        await adv_btn.click()
        await page.wait_for_timeout(500)

    # --- Fill date range ---
    # IDs contain dots so must use attribute selector, not # CSS shorthand.
    # Confirmed field IDs from live portal inspection 2026-04-18.
    await page.fill("input[id='caseCriteria.FileDateStart']", since_str)
    await page.fill("input[id='caseCriteria.FileDateEnd']", today_str)

    # --- Set case type to Estate ---
    # Tyler Odyssey uses a Typeahead autocomplete on caseCriteria.CaseType_input.
    # Dashboard/29 is already scoped to Probate division so this is belt-and-
    # suspenders filtering; skip silently if the autocomplete doesn't match.
    try:
        node_input = await page.query_selector("input[name='caseCriteria.CaseType_input']")
        if node_input:
            await node_input.fill("Estate")
            await page.wait_for_timeout(600)
            suggestion = await page.query_selector(
                ".tt-suggestion, .ui-autocomplete li, [role='option']"
            )
            if suggestion:
                await suggestion.click()
            await page.wait_for_timeout(300)
    except Exception as e:
        log.debug(f"Mahoning: CaseType autocomplete step skipped: {e}")

    # --- Wildcard search criteria (required by SmartSearch) ---
    try:
        await page.fill(
            "#caseCriteria_SearchCriteria, input[name='caseCriteria.SearchCriteria']", " "
        )
    except Exception:
        pass

    # --- Solve reCAPTCHA v2 ---
    log.info("Mahoning: requesting reCAPTCHA solve from 2captcha (15–45 s)…")
    token = await _twocaptcha_solve_recaptcha(_MAHONING_RECAPTCHA_SITEKEY, dashboard_url)

    # Inject the token into the hidden g-recaptcha-response textarea that Tyler
    # Odyssey reads on form submission. Some portal versions render the textarea
    # hidden; make it visible first so Playwright can set the value.
    await page.evaluate(
        """token => {
            const areas = document.querySelectorAll('[name="g-recaptcha-response"]');
            areas.forEach(el => {
                el.removeAttribute('disabled');
                el.style.display = '';
                el.value = token;
            });
        }""",
        token,
    )
    await page.wait_for_timeout(300)

    # --- Submit SmartSearch form ---
    # Intercept the JSON response from the SmartSearch POST so we can parse it
    # directly rather than scraping the rendered HTML table.
    search_response_data: list[dict] = []

    async def _capture_search_response(response):
        if "SmartSearch/SmartSearch" in response.url and response.status == 200:
            try:
                body = await response.json()
                search_response_data.extend(body if isinstance(body, list) else [body])
            except Exception:
                pass  # fall through to HTML parsing if JSON unavailable

    page.on("response", _capture_search_response)

    async with page.expect_navigation(wait_until="networkidle", timeout=60000):
        await page.click("#btnSSSubmit")

    page.remove_listener("response", _capture_search_response)

    # --- Parse results ---
    # Prefer the intercepted JSON; fall back to HTML table rows if empty.
    if search_response_data:
        raw_json = search_response_data[0]
        # Tyler Odyssey wraps results in SearchResults or a top-level list.
        case_list = raw_json.get("SearchResults") or (raw_json if isinstance(raw_json, list) else [])
        for item in case_list:
            case_type_raw = (item.get("NodeDesc") or item.get("CaseType") or "").strip()
            if case_type_raw and not any(
                t in case_type_raw.upper() for t in ("ESTATE", "PROBATE", "EST")
            ):
                continue
            results.append({
                "case_number": item.get("CaseNumber", ""),
                "owner_name":  item.get("StyleName") or item.get("CaseName") or "",
                "filing_date": item.get("FileDate") or item.get("FilingDate") or "",
                "case_type":   case_type_raw or "Estate",
                "_county":     "Mahoning",
                "_source_url": dashboard_url,
                "_address":    item.get("Address") or "",
            })
    else:
        # Fallback: parse the rendered HTML results table.
        # Tyler Odyssey table columns: Case Number | Style | File Date | Case Type | Status
        rows = await page.query_selector_all(
            "table#SmartSearchResults tr, table[id*='Results'] tr"
        )
        for row in rows:
            cells = await row.query_selector_all("td")
            if len(cells) < 3:
                continue
            texts = [await c.inner_text() for c in cells]
            if texts[0].strip().lower() in ("case number", "case no", ""):
                continue
            case_type_raw = texts[3].strip() if len(texts) > 3 else ""
            if case_type_raw and not any(
                t in case_type_raw.upper() for t in ("ESTATE", "PROBATE", "EST")
            ):
                continue
            results.append({
                "case_number": texts[0].strip(),
                "owner_name":  texts[1].strip() if len(texts) > 1 else "",
                "filing_date": texts[2].strip() if len(texts) > 2 else "",
                "case_type":   case_type_raw or "Estate",
                "_county":     "Mahoning",
                "_source_url": dashboard_url,
                "_address":    "",
            })

    # --- Pagination ---
    # Tyler Odyssey paginates via a "Next" link; repeat solve is NOT needed
    # because the pagination uses the existing session/search context.
    page_num = 1
    while True:
        next_link = await page.query_selector(
            "a[title='Next Page'], a[rel='next'], "
            f"a[href*='Page${page_num + 1}'], li.next a"
        )
        if not next_link:
            break
        await next_link.click()
        await page.wait_for_load_state("networkidle")
        await asyncio.sleep(random.uniform(2, 4))
        page_num += 1

        rows = await page.query_selector_all(
            "table#SmartSearchResults tr, table[id*='Results'] tr"
        )
        for row in rows:
            cells = await row.query_selector_all("td")
            if len(cells) < 3:
                continue
            texts = [await c.inner_text() for c in cells]
            if texts[0].strip().lower() in ("case number", "case no", ""):
                continue
            case_type_raw = texts[3].strip() if len(texts) > 3 else ""
            if case_type_raw and not any(
                t in case_type_raw.upper() for t in ("ESTATE", "PROBATE", "EST")
            ):
                continue
            results.append({
                "case_number": texts[0].strip(),
                "owner_name":  texts[1].strip() if len(texts) > 1 else "",
                "filing_date": texts[2].strip() if len(texts) > 2 else "",
                "case_type":   case_type_raw or "Estate",
                "_county":     "Mahoning",
                "_source_url": dashboard_url,
                "_address":    "",
            })

    log.info(f"Mahoning County: fetched {len(results)} probate filings since {since}")
    return results


# County scraper registry
COUNTY_SCRAPERS = {
    "cuyahoga": _fetch_cuyahoga,
    "lake": _fetch_lake,
    "mahoning": _fetch_mahoning,
}


# =============================================================================
# Parse — map raw scraped dict to raw_leads schema
# =============================================================================

def _parse_date(date_str: str) -> Optional[str]:
    """Parse MM/DD/YYYY or YYYY-MM-DD date strings to ISO format."""
    if not date_str:
        return None
    for fmt in ("%m/%d/%Y", "%Y-%m-%d", "%m-%d-%Y"):
        try:
            from datetime import datetime as dt
            return dt.strptime(date_str.strip(), fmt).date().isoformat()
        except ValueError:
            continue
    return None


def parse_filing(raw: dict, county: str, state: str) -> dict:
    """Map a raw scraped record to the raw_leads table schema.

    Extracts owner_name, filing_date, case_number (→ parcel_id placeholder),
    and source metadata. Property address is not available from probate search
    results — it is added during Step 1 enrichment via county auditor lookup.
    """
    filing_date = _parse_date(raw.get("filing_date", ""))

    # Filter non-Estate case types — some portals return mixed results
    case_type = raw.get("case_type", "").lower()
    if case_type and not any(t in case_type for t in ["estate", "probate", ""]):
        return {}

    # _address is the decedent's address from the portal results table.
    # It may be a mailing address, not the property address — enrichment will confirm.
    address_hint = raw.get("_address", "").strip() or None

    return {
        "owner_name": raw.get("owner_name", "").strip(),
        "property_address": address_hint,
        "parcel_id": None,              # populated by enrichment
        "filing_date": filing_date,
        "raw_data": raw,
        "source_type": "probate",
        "source_name": f"{county.title()} County Probate Court",
        "state": state,
        "county": county.title(),
        "verified_raw": False,
        "verified_enriched": False,
        "enriched": False,
    }


# =============================================================================
# Main agent — full 9-step pipeline
# =============================================================================

async def _run_async(county: str, state: str, lookback_days: int) -> None:
    """Async inner loop — runs Playwright and the full 9-step pipeline."""
    scraper = COUNTY_SCRAPERS.get(county.lower())
    if not scraper:
        log.error(f"No scraper registered for county: {county}")
        return

    since = date.today() - timedelta(days=lookback_days)
    log.info(f"Probate agent starting — {county.title()} County, {state} (since {since})")

    sbr_ws = os.environ.get("BRIGHTDATA_WS_URL", "").strip()

    async with async_playwright() as pw:
        # Lake County requires BrightData Scraping Browser to bypass Cloudflare.
        # If BRIGHTDATA_SBR_WS is set, connect via CDP; otherwise fall through to
        # the standard browser (which will hit CloudflareBlockedError and be caught).
        if county.lower() == "lake" and sbr_ws:
            log.info("Lake County: connecting via BrightData Scraping Browser (CDP)")
            try:
                browser = await pw.chromium.connect_over_cdp(sbr_ws, timeout=30000)
            except Exception as e:
                log.error(f"Lake County: Scraping Browser CDP connect failed: {e}")
                return
            context = browser.contexts[0] if browser.contexts else await browser.new_context()
        else:
            browser = await pw.chromium.launch(headless=True)
            context_kwargs: dict = {
                "user_agent": (
                    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                    "AppleWebKit/537.36 (KHTML, like Gecko) "
                    "Chrome/120.0.0.0 Safari/537.36"
                )
            }
            if county.lower() == "lake":
                # No Scraping Browser configured — will attempt anyway; CloudflareBlockedError
                # will be raised inside _fetch_lake and caught below.
                log.warning(
                    "Lake County: BRIGHTDATA_SBR_WS not set — attempting without Scraping Browser. "
                    "Expected to fail; set BRIGHTDATA_SBR_WS in .env to fix."
                )
            context = await browser.new_context(**context_kwargs)

        page = await context.new_page()

        # 1. FETCH
        try:
            raw_filings = await scraper(page, since)
        except (CloudflareBlockedError, Exception) as e:
            # BrightData policy blocks and Cloudflare hard-blocks are both known
            # infrastructure gaps — not selector bugs. Catch them silently, flag
            # the source, and alert the owner. Do NOT trigger self_healer.
            err_str = str(e)
            is_infra_block = (
                isinstance(e, CloudflareBlockedError)
                or "classified as Government" in err_str
                or "brightdata.com/proxy-networks" in err_str
                or "proxy_error" in err_str
            )
            if is_infra_block:
                log.error(f"FETCH blocked for {county} (proxy/CF policy): {e}")
                try:
                    get_client().table("sources").update({"blocked": True}).eq(
                        "source_name", "Lake County Probate Court"
                    ).execute()
                except Exception:
                    pass
                from routing.notify import send_sms
                send_sms(
                    "[SOURCE ALERT] Lake County probate blocked (BrightData policy). "
                    "Leads paused. Contact BrightData support to whitelist "
                    "phoenix.lakecountyohio.gov, or switch to Zyte/Oxylabs."
                )
            else:
                log.error(f"FETCH failed for {county}: {e}")
                from maintenance.self_healer import handle_failure
                handle_failure(f"probate_{county}", err_str)
            await browser.close()
            return

        await browser.close()

    if not raw_filings:
        log.warning(f"No filings returned for {county} — check portal or selectors")
        from maintenance.self_healer import handle_failure
        handle_failure(f"probate_{county}", "Zero records returned")
        return

    log.info(f"Processing {len(raw_filings)} raw filings for {county}")
    new_records = 0
    tier_a_count = 0
    tier_b_count = 0

    for raw in raw_filings:
        try:
            # 2. PARSE
            record = parse_filing(raw, county, state)
            if not record or not record.get("owner_name"):
                continue

            # 3. DEDUPE
            if is_duplicate(
                county=county.title(),
                source_type="probate",
                parcel_id=record.get("parcel_id"),
                property_address=record.get("property_address"),
                owner_name=record.get("owner_name"),
            ):
                log.debug(f"Duplicate: {record.get('owner_name')} — skipping")
                continue

            # 4. STORE
            stored = insert_row("raw_leads", record)
            lead_id = stored.get("id")
            if not lead_id:
                log.error(f"Insert returned no ID for {record.get('owner_name')}")
                continue
            new_records += 1

            # 5. VERIFY (Gate 1)
            passed_gate1 = verify_raw_record(lead_id)
            if not passed_gate1:
                log.debug(f"Lead {lead_id} failed Gate 1 — skipping enrichment")
                continue

            # 6. ENRICH (waterfall: public sources → Skip Sherpa → Skip Matrix flag)
            enrich_lead(lead_id)

            # 7. VERIFY (Gate 2) — called inside enrich_lead after enrichment
            # Re-fetch to get current verified_enriched state
            refreshed = (
                get_client()
                .table("raw_leads")
                .select("*")
                .eq("id", lead_id)
                .single()
                .execute()
                .data
            )
            if not refreshed or not refreshed.get("verified_enriched"):
                log.debug(f"Lead {lead_id} failed Gate 2 — skipping scoring")
                continue

            # 8. SCORE
            result = score_lead(refreshed)
            update_row("raw_leads", lead_id, {
                **result,
                "scored_at": "now()",
            })
            log.info(
                f"Scored: {record.get('owner_name')} | "
                f"distress={result['distress_score']} deal={result['deal_score']} "
                f"total={result['score']} tier={result['tier']}"
            )

            # 9. ROUTE
            if result["tier"] == "A":
                tier_a_count += 1
                route_lead(lead_id, result["tier"])
            elif result["tier"] == "B":
                tier_b_count += 1
                route_lead(lead_id, result["tier"])

        except Exception as e:
            log.error(f"Error processing filing for {county}: {e}")
            continue

    log.info(
        f"Probate agent complete — {county.title()} County | "
        f"{new_records} new records stored | {tier_a_count} Tier A, {tier_b_count} Tier B"
    )


# =============================================================================
# Excel file ingest — Mahoning County probate data
#
# File received from VA: "Lake ounty Mahoning_ Ohio.xlsx", sheet "MAHONING-100"
# Column layout (confirmed from file received 2026-04):
#
# Col 0: Case Number (row counter — not a real case number)
# Col 1: Deceased name   → owner_name  (format: "Last, First M.")
# Col 2: Deceased address → street portion of property address
# Col 3: De city, state, zip code → "City,\xa0ST\xa0XXXXX" format
# Col 4: PR name          → raw_data.pr_name
# Col 5: PR Address       → raw_data.pr_address (PR mailing address)
# Col 6: PR city, state, zip code → raw_data.pr_city_state_zip
# Col 7: Phone number     → raw_data.pr_phone (pre-skip-traced — already enriched)
#
# No filing date in the file — filing_date is set to ingest date (today).
# Dedup key: owner_name + property_address (no parcel ID in this file).
# =============================================================================

def _clean_city_state_zip(raw: str) -> str:
    """Normalize city/state/zip field from Mahoning file.

    Input: 'Youngstown,\\xa0OH\\xa044515' (non-breaking spaces around state code)
    Output: 'Youngstown, OH 44515'
    """
    if not raw:
        return ""
    cleaned = str(raw).replace("\xa0", " ").strip()
    # Ensure comma after city if missing
    parts = cleaned.split(",", 1)
    if len(parts) == 2:
        return f"{parts[0].strip()}, {parts[1].strip()}"
    return cleaned


def ingest_mahoning_probate_file(filepath: str, state: str = "OH") -> None:
    """Ingest a Mahoning County probate Excel file through the full pipeline.

    Reads the 'MAHONING-100' sheet, maps columns to the raw_leads schema,
    and runs each record through the standard 9-step pipeline.

    Called manually:
        python agents/probate_agent.py --ingest-file <path>

    The PR phone number (col 7) is stored in raw_data.pr_phone and serves as
    a pre-skip-traced contact — Gate 2 enrichment checks for it automatically.
    """
    import openpyxl

    log.info(f"Ingesting Mahoning County probate file: {filepath}")

    wb = openpyxl.load_workbook(filepath)
    if "MAHONING-100" not in wb.sheetnames:
        # Fall back to first sheet
        ws = wb.active
        log.warning(f"'MAHONING-100' sheet not found — using '{ws.title}'")
    else:
        ws = wb["MAHONING-100"]

    raw_records: list[dict] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        deceased_name = str(row[1]).strip() if row[1] else ""
        if not deceased_name or deceased_name.lower() == "none":
            continue

        street = str(row[2]).strip() if row[2] else ""
        city_state_zip = _clean_city_state_zip(str(row[3]) if row[3] else "")
        property_address = f"{street}, {city_state_zip}" if street and city_state_zip else (street or city_state_zip or None)

        raw_records.append({
            "owner_name": deceased_name,
            "_address": property_address,   # parse_filing() reads "_address", not "property_address"
            "filing_date": date.today().isoformat(),
            "raw_data": {
                "pr_name": str(row[4]).strip() if row[4] else "",
                "pr_address": str(row[5]).strip() if row[5] else "",
                "pr_city_state_zip": _clean_city_state_zip(str(row[6]) if row[6] else ""),
                "pr_phone": str(row[7]).strip() if row[7] else "",
                "source_file": filepath,
            },
        })

    log.info(f"Mahoning probate file: {len(raw_records)} records parsed")

    if not raw_records:
        log.warning("No records parsed from Mahoning probate file — check file format")
        return

    client = get_client()
    new_records = 0
    tier_a_count = 0
    tier_b_count = 0

    for raw in raw_records:
        try:
            record = parse_filing(raw, "mahoning", state)
            if not record or not record.get("owner_name"):
                continue

            if is_duplicate(
                county="Mahoning",
                source_type="probate",
                parcel_id=None,
                property_address=record.get("property_address"),
                owner_name=record.get("owner_name"),
            ):
                continue

            stored = insert_row("raw_leads", record)
            lead_id = stored.get("id")
            if not lead_id:
                log.error(f"Insert returned no ID for {record.get('owner_name')}")
                continue
            new_records += 1

            if not verify_raw_record(lead_id):
                log.debug(f"Lead {lead_id} failed Gate 1 — skipping enrichment")
                continue

            enrich_lead(lead_id)

            refreshed = (
                client.table("raw_leads")
                .select("*")
                .eq("id", lead_id)
                .single()
                .execute()
                .data
            )
            if not refreshed or not refreshed.get("verified_enriched"):
                log.debug(f"Lead {lead_id} failed Gate 2 — skipping scoring")
                continue

            result = score_lead(refreshed)
            update_row("raw_leads", lead_id, {**result, "scored_at": "now()"})
            log.info(
                f"Scored: {record.get('owner_name')} | "
                f"distress={result['distress_score']} deal={result['deal_score']} "
                f"total={result['score']} tier={result['tier']}"
            )

            if result["tier"] == "A":
                tier_a_count += 1
                route_lead(lead_id, result["tier"])
            elif result["tier"] == "B":
                tier_b_count += 1
                route_lead(lead_id, result["tier"])

        except Exception as e:
            log.error(f"Error processing Mahoning probate record: {e}")
            continue

    if tier_a_count > 0:
        send_sms(
            f"[Lead Intel] Mahoning probate file: {new_records} new leads — "
            f"{tier_a_count} Tier A, {tier_b_count} Tier B."
        )

    log.info(
        f"Mahoning probate file ingest complete — {new_records} new records stored "
        f"({tier_a_count} Tier A, {tier_b_count} Tier B)"
    )


def run(county: str, state: str = "OH", lookback_days: int = DEFAULT_LOOKBACK_DAYS) -> None:
    """Run the probate agent for one county (synchronous entry point)."""
    asyncio.run(_run_async(county, state, lookback_days))


# =============================================================================
# CLI
# =============================================================================

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Probate agent — Ohio POC")
    parser.add_argument("--county", help="County name (cuyahoga | lake | mahoning)")
    parser.add_argument("--state", default="OH", help="State code (default: OH)")
    parser.add_argument("--days", type=int, default=DEFAULT_LOOKBACK_DAYS,
                        help=f"Lookback days (default: {DEFAULT_LOOKBACK_DAYS})")
    parser.add_argument("--all-counties", action="store_true",
                        help="Run all three Ohio POC counties")
    parser.add_argument(
        "--ingest-file",
        metavar="PATH",
        help="Ingest a Mahoning County probate Excel file directly (skips scraping)",
    )
    args = parser.parse_args()

    if args.ingest_file:
        ingest_mahoning_probate_file(args.ingest_file, args.state)
    elif args.all_counties:
        for county in OHIO_COUNTIES:
            run(county, args.state, args.days)
    elif args.county:
        run(args.county.lower(), args.state, args.days)
    else:
        parser.print_help()
